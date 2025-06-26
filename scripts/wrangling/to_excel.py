import polars as pl
import requests
import zipfile
import io
from openpyxl import Workbook
from markdownify import markdownify as md
import click
from lib import data_url


@click.command()
@click.argument(
    "input_path", type=click.Path(exists=True), required=True, metavar="INPUT_FILE"
)
@click.argument(
    "output_folder",
    required=True,
    type=click.Path(file_okay=False, writable=True),
    metavar="OUTPUT_FOLDER",
)
def process(input_path, output_folder):
    response = requests.get(data_url)
    response.raise_for_status()

    with zipfile.ZipFile(io.BytesIO(response.content)) as z:
        file_name = "min_data_format.csv"
        with z.open(file_name) as f:
            df_all = pl.read_csv(f)

    df_labeled = pl.read_json(input_path)

    workbook = Workbook()

    del workbook["Sheet"]

    sheet_all = workbook.create_sheet(title="all_data")

    for col_num, column_name in enumerate(df_all.columns, 1):
        sheet_all.cell(row=1, column=col_num, value=column_name)

    for row_num, row in enumerate(df_all.rows(), 2):
        for col_num, value in enumerate(row, 1):
            sheet_all.cell(row=row_num, column=col_num, value=str(value))

    sheet_labeled = workbook.create_sheet(title="positive_examples")

    format_cols = ["description", "more_info", "legal_basis"]

    print("formatting columns...")
    for col in format_cols:
        df_labeled = df_labeled.with_columns(
            pl.col(col).map_elements(md, return_dtype=pl.String)
        )

    df_labeled = df_labeled.filter(pl.col("label") == 1)

    for col_num, column_name in enumerate(df_labeled.columns, 1):
        sheet_labeled.cell(row=1, column=col_num, value=column_name)

    # write labeled data to the second sheet
    for row_num, row in enumerate(df_labeled.rows(), 2):
        for col_num, value in enumerate(row, 1):
            sheet_labeled.cell(row=row_num, column=col_num, value=str(value))

    workbook.save(f"{output_folder}/data.xlsx")


if __name__ == "__main__":
    process()
